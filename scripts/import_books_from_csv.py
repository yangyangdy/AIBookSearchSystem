"""从 CSV / XLSX 导入数据到 book_info_table"""
import sys
import csv
import argparse
from pathlib import Path
from typing import List, Dict, Any

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from sqlalchemy import text
from loguru import logger
from src.core.mysql_client import MySQLClient
from src.utils.config import get_settings


# 表 book_info_table 的列（与建表一致；不含 id，使用表自增）
TABLE_COLUMNS = [
    "sku", "isbn", "create_time", "update_time", "binding_layout", "row_number",
    "name", "reference_name", "level", "inventory", "min_age", "max_age", "age_group",
    "market_price", "shendong_price", "author", "publishing_house", "shelves",
    "image_link", "cover_link", "inside_pages_link", "inside_pages_link2", "back_cover",
    "tag", "record", "weight", "size", "synopsis", "import_record_id", "page_number",
    "publishing_time", "sale_point", "isbn2",
]


def _normalize_value(val: Any) -> Any:
    """空字符串转 None，数值字符串尝试转 int/float"""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    s = val.strip() if isinstance(val, str) else val
    if isinstance(s, str):
        try:
            if "." in s:
                return float(s)
            return int(s)
        except ValueError:
            pass
    return s if isinstance(val, str) else val


def read_csv(path: Path, encoding: str = "utf-8") -> List[Dict[str, Any]]:
    """读取 CSV，第一行为列名，返回字典列表。"""
    rows = []
    with open(path, "r", encoding=encoding, newline="", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({k.strip(): v for k, v in row.items() if k})
    return rows


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    """读取 XLSX，第一行为列名，返回字典列表。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    # 第一行作为列名
    headers = []
    for cell in ws[1]:
        h = cell.value
        headers.append(str(h).strip() if h is not None else "")
    headers = [h or f"_col{i}" for i, h in enumerate(headers)]
    rows = []
    for row in ws.iter_rows(min_row=2):
        rec = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                rec[headers[i]] = cell.value
        if rec:
            rows.append(rec)
    wb.close()
    return rows


def read_file(path: Path, encoding: str = "utf-8") -> List[Dict[str, Any]]:
    """根据扩展名读取 CSV 或 XLSX，第一行为列名，返回字典列表。"""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix in (".csv", ".txt"):
        return read_csv(path, encoding=encoding)
    raise ValueError(f"不支持的文件格式: {suffix}，请使用 .csv 或 .xlsx")


def insert_batch(
    client: MySQLClient,
    table_name: str,
    rows: List[Dict[str, Any]],
    columns: List[str],
    batch_size: int = 500,
) -> int:
    """批量插入，返回插入行数。"""
    if not rows:
        return 0
    session = client.get_session()
    total = 0
    try:
        for i in range(0, len(rows), batch_size):
            batch = rows[i : i + batch_size]
            # 构建多行 INSERT: (a,b,c) VALUES (:a_0,:b_0,:c_0), (:a_1,:b_1,:c_1), ...
            placeholders = []
            params = {}
            for idx, row in enumerate(batch):
                tokens = []
                for c in columns:
                    key = f"{c}_{idx}"
                    params[key] = _normalize_value(row.get(c))
                    tokens.append(f":{key}")
                placeholders.append("(" + ", ".join(tokens) + ")")
            col_list = ", ".join(columns)
            values_clause = ", ".join(placeholders)
            sql = text(f"INSERT INTO {table_name} ({col_list}) VALUES {values_clause}")
            session.execute(sql, params)
            session.commit()
            total += len(batch)
            logger.info(f"已插入 {total}/{len(rows)} 条")
    except Exception as e:
        session.rollback()
        logger.error(f"插入失败: {e}")
        raise
    finally:
        session.close()
    return total


def main():
    parser = argparse.ArgumentParser(
        description="从 CSV 或 XLSX 导入数据到 book_info_table"
    )
    parser.add_argument(
        "file_path",
        type=str,
        help="数据文件路径（.csv 或 .xlsx，第一行为列名）",
    )
    parser.add_argument("--batch-size", type=int, default=500, help="每批插入条数")
    parser.add_argument("--encoding", type=str, default="utf-8", help="CSV 编码（仅对 CSV 生效）")
    parser.add_argument(
        "--table",
        type=str,
        default=None,
        help="表名（默认使用 config 中的 table_name）",
    )
    args = parser.parse_args()

    path = Path(args.file_path)
    if not path.exists():
        logger.error(f"文件不存在: {path}")
        if "\\" not in args.file_path and ("Users" in args.file_path or "Downloads" in args.file_path):
            logger.info(
                "提示: 在 bash 中反斜杠会被转义，请用引号包裹路径或使用正斜杠，例如："
            )
            logger.info('  python scripts/import_books_from_csv.py "D:\\Users\\...\\文件.xlsx"')
            logger.info("  或  python scripts/import_books_from_csv.py D:/Users/.../文件.xlsx")
        sys.exit(1)

    settings = get_settings()
    table_name = args.table or settings.mysql.table_name
    client = MySQLClient()

    try:
        rows = read_file(path, encoding=args.encoding)
        logger.info(f"读取 {path.suffix.upper()} 共 {len(rows)} 条，表: {table_name}")
        if not rows:
            logger.warning("文件无数据")
            return
        sample = rows[0]
        available = [c for c in TABLE_COLUMNS if c in sample]
        inserted = insert_batch(
            client, table_name, rows, available, batch_size=args.batch_size
        )
        logger.info(f"导入完成，共插入 {inserted} 条")
    finally:
        client.close()


if __name__ == "__main__":
    main()
